import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from google.oauth2.service_account import Credentials
import gspread
import os

print(os.path.isfile("credenciais.json"))
print(os.path.getsize("credenciais.json"))

scopes = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

credentials = Credentials.from_service_account_file("credenciais.json", scopes=scopes)
client = gspread.authorize(credentials)

spreadsheet = client.open("MONITORAMENTO JTD")

hoje = datetime.today()
nome_aba = hoje.strftime("%d/%m/%Y")
sheet = spreadsheet.worksheet(nome_aba)

data = sheet.get_all_records()
df = pd.DataFrame(data)

ontem = hoje - timedelta(days=1)
data_ontem = ontem.strftime("%d/%m/%Y")

df_filtrado = df[df['Data_expedicao'] == data_ontem]

if df_filtrado.empty:
    print("Nenhum dado encontrado para ontem.")
else:
    caminho_arquivo_destino = r"\\server\JTDTRANSPORTES2\ANALITCS\ACOMPANHAMENTO DE CARGAS\CARGAS DIRETA - MONDIAL.xlsx"

    book = load_workbook(caminho_arquivo_destino)
    sheet_name = 'Cargas'
    sheet_destino = book[sheet_name]

    ultima_linha = sheet_destino.max_row

    with pd.ExcelWriter(caminho_arquivo_destino, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df_filtrado.to_excel(writer, index=False, header=False, sheet_name=sheet_name, startrow=ultima_linha)

    print(f"{len(df_filtrado)} linha(s) adicionada(s) na planilha de destino.")